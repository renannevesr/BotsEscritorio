import time
import pandas as pd
from selenium.webdriver.common.by import By
from selenium import webdriver
import base64
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import pdfkit
import os
import shutil
import re

url = 'https://www.jfrs.jus.br/projefweb/'
list_b = 'Auditor.csv'
df_b = pd.read_csv(list_b, sep =';')
print(df_b.head())
rows = df_b.shape[0]
df_b['Polo Passivo'] = df_b['Polo Passivo'].fillna('INSS/UNIÃO')
options = Options()
download_dir = r"C:\Users\Renan\Desktop\projefweb\pdf"
destination_directory = r"C:\Users\Renan\Desktop\projefweb\completo"
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "plugins.always_open_pdf_externally": True
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=options)
driver.get(url)




mapping = {
    ('procurador', '1/1'): '64692f59',
    ('procurador', '1/2'): '3003a4ff',
    ('procurador', '1/4'): '0ba80a01',
    ('procurador', '12/36'): '745f57dd',
    ('auditor', '1/1'): '2a735e7e',
    ('auditor', '1/2'): '6b74c5a0',
    ('auditor', '1/3'): '05f9f561',
    ('auditor', '1/4'): '4ac50d19',
    ('auditor', '1/6'): '2580f5bb',
    ('auditor', '1/10'): '76a6527b',
    ('auditor', '1/16'): '7392bf29'
}

def map_values(row):
    key = ('auditor',row['COTA'])
    return mapping.get(key)  

df_b['id'] = df_b.apply(map_values, axis=1)

def organization(nome,polo,parts_cota):
    print("entrou")

    polo_passivo_folder = os.path.join(destination_directory, polo)
    cota_parte_folder = os.path.join(polo_passivo_folder,parts_cota)

    os.makedirs(cota_parte_folder, exist_ok=True)

    # Encontra o arquivo PDF correspondente
    nome_fragment = nome.replace(" ", "-").upper()
    pattern = re.compile(re.escape(nome_fragment), re.IGNORECASE)
    for pdf_file in os.listdir(download_dir):
        if pattern.search(pdf_file):
            source = os.path.join(download_dir, pdf_file)
            destination = os.path.join(cota_parte_folder, pdf_file)
            shutil.move(source, destination)
            print(f'Movido: {pdf_file} para {destination}')
            break

def calc(nome,number):
    file_path = f'Modelo Auditor - 1_{number}.xlsx'
    new_file_path = 'Modelo_Auditor_Editado.xlsx'
    xls = pd.ExcelFile(file_path)
    sheet_name = xls.sheet_names[0]  
    df = xls.parse(sheet_name)
    df.iloc[3, 1] = nome
    with pd.ExcelWriter(new_file_path) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        
def pdf (nome, number):
    xls_file_path = 'Modelo_Auditor_Editado.xlsx'
    html_file_path = 'Modelo_Auditor_Editado.html'
    nome_arq = f"{nome}.pdf"
    pdf_file_path = os.path.join(number, nome_arq)
    wb = load_workbook(xls_file_path)
    ws = wb.active
    html_content = '<html><head><style>table, th, td { border: 1px solid black; border-collapse: collapse; padding: 5px; }</style></head><body>'
    html_content += '<table>'
    for row in ws.iter_rows():
        html_content += '<tr>'
        for cell in row:
            html_content += f'<td>{cell.value}</td>'
        html_content += '</tr>'
    html_content += '</table></body></html>'
    with open(html_file_path, 'w') as f:
        f.write(html_content)
    pdfkit.from_file(html_file_path, pdf_file_path)

    
    
for i in range(rows):
    driver.get(url)
    janela_original = driver.current_window_handle
    indice = 0
    idx = 0
    id = df_b.loc[i, 'id']
    nome = df_b.loc[i, 'Nome']
    cota = df_b.loc[i, 'COTA']
    polo = df_b.loc[i, 'Polo Passivo']
    parts_cota = cota.replace("/", ".")
    driver.find_element("xpath", '//*[@id="ext-gen316"]').click()
    time.sleep(1)
    inputs = driver.find_elements(By.TAG_NAME, "input")
    time.sleep(2)
    inputs[55].send_keys(id)
    inputs[55].send_keys(Keys.ENTER)
    time.sleep(2)
    driver.find_element("xpath", '//*[@id="ext-gen73"]').clear()
    driver.find_element("xpath", '//*[@id="ext-gen73"]').send_keys(nome)
    driver.find_element("xpath", '//*[@id="ext-gen75"]').clear()
    driver.find_element("xpath", '//*[@id="ext-gen75"]').send_keys(polo)
    time.sleep(1)
    spans = driver.find_elements(By.TAG_NAME, "span")
    spans[12].click()
    driver.find_element("xpath", '//*[@id="ext-gen411"]').click()
    time.sleep(1)
    organization(nome,polo,parts_cota)

    #calc(nome,number)
    time.sleep(1)
   # pdf (nome, number)
    time.sleep(1)
